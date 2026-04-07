"""
Export Carousell listings from microsearch.py Excel output → src/data/listings.ts

Usage:
  python scripts/export_listings.py

This reads the latest Excel from microsearch.py and overwrites
src/data/listings.ts with real live listings data.

Run this after microsearch.py has completed a scrape.
"""

import re
import json
from pathlib import Path
from datetime import datetime, date

EXCEL_PATH = Path(r"C:\Users\benbu\OneDrive\AI Projects\Microbrand watch search\Microbrand Watch Search.xlsx")
OUTPUT_TS  = Path(__file__).parent.parent / "src" / "data" / "listings.ts"

# HKD → USD rough conversion (update as needed)
HKD_TO_USD = 0.128

# Known microbrand names — only listings whose Brand column matches one of these
# are included. This prevents luxury watch spam (Rolex, etc.) from appearing.
KNOWN_BRANDS = {
    'Aevig', 'Akrone', 'AnOrdain', 'Aquatico', 'Aragon', 'Astor + Banks',
    'Atelier Wen', 'Autodromo', 'Baltic', 'Beaubleu', 'Bertucci', 'Boldr',
    'Borealis', 'Bravur', 'Brew', 'Briston', 'Christopher Ward', 'Circula',
    'Clemence', 'Clemens', 'Code41', 'Damasko', 'Dan Henry', 'Dekla',
    'Dennison', 'Depancel', 'Direnzo', 'Dufrane', 'Elliot Brown', 'Enoksen',
    'Eza', 'Farer', 'Fears', 'Formex', 'Furlan Marri', 'Ginault',
    'Gruppo Gamma', 'Halios', 'Hanhart', 'Helm', 'Henry Archer', 'Hoffman',
    'Ikepod', 'Isotope', 'Jack Mason', 'James Brand', 'Jan Sarnowski',
    'Knot', 'Kuoe', 'Kurono Tokyo', 'Laco', 'Lorier', 'Maen', 'Ming',
    'Monta', 'Mr Jones Watches', 'Muhle Glashutte', 'Nivada Grenchen',
    'Nodus', 'Nomos', 'Oak & Oscar', 'Ophion', 'Orion', 'Paulin', 'Pelton',
    'Phoibos', 'Pitzmann', 'Praesidus', 'Raven Watches', 'Redwood',
    'Revelot', 'RZE', 'Sangin Instruments', 'Serica', 'Smiths', 'Squale',
    'Steinhart', 'Straum', 'Studio Underd0g', 'Traska', 'Tsao Baltimore',
    'Tuseno', 'Undone', 'Unimatic', 'Vaer', 'Vario', 'Ventus', 'Vertex',
    'Wise', 'Wolbrook', 'Yema', 'Zelos', 'Zodiac', 'Loresum',
}


def hkd_to_usd(price_str: str) -> int | None:
    """Parse a price string like 'HK$1,200' or '1200' and convert to USD."""
    if not price_str or str(price_str).strip() in ('', 'None', 'nan'):
        return None
    cleaned = re.sub(r'[^0-9.]', '', str(price_str))
    if not cleaned:
        return None
    try:
        return round(float(cleaned) * HKD_TO_USD)
    except ValueError:
        return None


def parse_platform(url: str) -> str:
    if 'carousell' in str(url).lower():
        return 'Carousell'
    if 'ebay' in str(url).lower():
        return 'eBay'
    return 'Other'


def days_old_to_date(days_old) -> str:
    """Convert days_old integer to an ISO date string."""
    try:
        d = int(days_old)
        result = date.today().toordinal() - d
        return date.fromordinal(result).isoformat()
    except (TypeError, ValueError):
        return date.today().isoformat()


def slug_brand(name: str) -> str:
    return name.strip()


def main():
    try:
        import openpyxl
    except ImportError:
        print("openpyxl not installed. Run: pip install openpyxl")
        return

    if not EXCEL_PATH.exists():
        print(f"Excel file not found: {EXCEL_PATH}")
        print("Run microsearch.py first to generate the Excel output.")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # Read header row
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    print(f"Columns: {headers}")

    # Map column names to indices (1-based)
    col = {name: idx + 1 for idx, name in enumerate(headers) if name}

    rows = []
    for r in range(2, ws.max_row + 1):
        brand    = ws.cell(r, col.get('Brand', 1)).value
        title    = ws.cell(r, col.get('Title', 2)).value
        price    = ws.cell(r, col.get('Price (HKD)', 3)).value
        days_old = ws.cell(r, col.get('Days Old', 5)).value
        # URLs are stored as hyperlinks, not cell values
        link_cell = ws.cell(r, col.get('Link', 6))
        url = getattr(link_cell.hyperlink, 'target', None) or link_cell.value

        if not brand or not title:
            continue

        # Skip rows where URL is missing or literally "Link" (header bleed)
        url_str = str(url) if url else ''
        if not url_str or url_str.strip().lower() in ('link', 'none', ''):
            continue

        # Only include known microbrand names
        brand_str = str(brand).strip()
        if brand_str not in KNOWN_BRANDS:
            continue

        # Skip titles that are mostly non-ASCII (spam/junk listings)
        title_str = str(title).strip()
        non_ascii = sum(1 for c in title_str if ord(c) > 127)
        if non_ascii > len(title_str) * 0.4:
            continue

        usd_price = hkd_to_usd(price)
        if usd_price is None or usd_price <= 0 or usd_price > 15000:
            continue  # skip free/no-price listings and absurd prices

        rows.append({
            'brand':    slug_brand(str(brand)),
            'title':    str(title).strip(),
            'price':    usd_price,
            'days_old': days_old,
            'url':      str(url) if url else 'https://carousell.com.hk',
        })

    if not rows:
        print("No valid listings found in Excel. Check the file and column names.")
        return

    # Sort by newest first
    rows.sort(key=lambda x: x['days_old'] if isinstance(x['days_old'], (int, float)) else 999)

    # Build TypeScript source
    ts_lines = [
        "// AUTO-GENERATED by scripts/export_listings.py — do not edit manually",
        f"// Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        "// Source: Microbrand Watch Search.xlsx (microsearch.py output)",
        "",
        "export interface Listing {",
        "  id: string",
        "  brand: string",
        "  title: string",
        "  price: number",
        "  currency: string",
        "  condition: 'New' | 'Like New' | 'Good' | 'Fair'",
        "  platform: 'Carousell' | 'eBay' | 'Other'",
        "  location: string",
        "  url: string",
        "  postedAt: string",
        "  imageUrl?: string",
        "}",
        "",
        "export const listings: Listing[] = [",
    ]

    for i, row in enumerate(rows):
        platform = parse_platform(row['url'])
        location = 'Hong Kong' if 'carousell.com.hk' in row['url'] else 'Unknown'
        posted   = days_old_to_date(row['days_old'])

        # Condition heuristic: assume Good unless price looks high (>80% of new)
        condition = 'Good'

        entry = (
            f"  {{\n"
            f"    id: '{i + 1}',\n"
            f"    brand: {json.dumps(row['brand'])},\n"
            f"    title: {json.dumps(row['title'])},\n"
            f"    price: {row['price']},\n"
            f"    currency: 'USD',\n"
            f"    condition: '{condition}',\n"
            f"    platform: '{platform}',\n"
            f"    location: '{location}',\n"
            f"    url: {json.dumps(row['url'])},\n"
            f"    postedAt: '{posted}',\n"
            f"  }},"
        )
        ts_lines.append(entry)

    ts_lines += [
        "]",
        "",
        "export const listingBrands = [...new Set(listings.map(l => l.brand))].sort()",
        "export const listingPlatforms = [...new Set(listings.map(l => l.platform))].sort()",
        "export const listingConditions = ['New', 'Like New', 'Good', 'Fair'] as const",
        "",
    ]

    OUTPUT_TS.write_text('\n'.join(ts_lines), encoding='utf-8')
    print(f"\nExported {len(rows)} listings -> {OUTPUT_TS}")
    print(f"Brands found: {sorted(set(r['brand'] for r in rows))}")


if __name__ == '__main__':
    main()
